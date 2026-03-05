# backend/app/export.py
import pandas as pd
import io
from datetime import datetime
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ScheduleExporter:
    def __init__(self, db):
        self.db = db
        self.days_ru = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
        self.time_slots = [
            {"num": 1, "time": "08:40-10:00"},
            {"num": 2, "time": "10:10-11:30"},
            {"num": 3, "time": "12:00-13:20"},
            {"num": 4, "time": "13:30-14:50"},
            {"num": 5, "time": "15:00-16:20"},
            {"num": 6, "time": "16:30-17:50"}
        ]
    
    def create_schedule_table(self, start_date, end_date, group_id=None):
        """
        Создать структуру таблицы для отображения на сайте
        """
        try:
            from .models import Schedule, Group
            
            logger.info(f"Creating schedule table for {start_date} - {end_date}, group_id={group_id}")
            
            # Получаем расписание
            query = self.db.query(Schedule).filter(
                Schedule.date >= start_date,
                Schedule.date <= end_date
            ).order_by(Schedule.date, Schedule.start_time)
            
            if group_id:
                query = query.filter(Schedule.group_id == group_id)
            
            schedules = query.all()
            
            logger.info(f"Found {len(schedules)} schedule entries")
            
            if not schedules:
                return {
                    'days': [],
                    'rpo_groups': [],
                    'kgid_groups': []
                }
            
            # Группируем по дням
            days_dict = {}
            rpo_groups = set()
            kgid_groups = set()
            
            for s in schedules:
                date_str = s.date.strftime('%d.%m.%Y')
                if date_str not in days_dict:
                    day_name = self.days_ru[s.date.weekday()] if s.date.weekday() < len(self.days_ru) else ''
                    days_dict[date_str] = {
                        'date': date_str,
                        'day': day_name,
                        'pairs': {i: {'rpo': [], 'kgid': []} for i in range(1, 7)}
                    }
                
                # Определяем номер пары
                pair_num = 1
                for i, slot in enumerate(self.time_slots, 1):
                    if s.start_time in slot["time"]:
                        pair_num = i
                        break
                
                # Определяем тип группы
                group_type = 'rpo' if 'рпо' in s.group.name.lower() else 'kgid'
                if group_type == 'rpo':
                    rpo_groups.add(s.group.name)
                else:
                    kgid_groups.add(s.group.name)
                
                days_dict[date_str]['pairs'][pair_num][group_type].append({
                    'subject': s.subject.name if s.subject else '',
                    'teacher': s.teacher.name if s.teacher else '',
                    'room': s.room,
                    'group': s.group.name
                })
            
            # Формируем результат
            result = {
                'days': [],
                'rpo_groups': list(rpo_groups),
                'kgid_groups': list(kgid_groups)
            }
            
            # Сортируем даты
            sorted_dates = sorted(days_dict.keys(), key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
            
            for date_str in sorted_dates:
                day_info = days_dict[date_str]
                day_data = {
                    'date': date_str,
                    'day': day_info['day'],
                    'pairs': []
                }
                
                for pair_num in range(1, 7):
                    pair_data = {
                        'num': pair_num,
                        'time': self.time_slots[pair_num-1]["time"],
                        'rpo': day_info['pairs'][pair_num]['rpo'],
                        'kgid': day_info['pairs'][pair_num]['kgid']
                    }
                    day_data['pairs'].append(pair_data)
                
                result['days'].append(day_data)
            
            logger.info(f"Created schedule table with {len(result['days'])} days")
            return result
            
        except Exception as e:
            logger.error(f"Error creating schedule table: {e}")
            import traceback
            traceback.print_exc()
            return {
                'days': [],
                'rpo_groups': [],
                'kgid_groups': []
            }
    
    def export_to_excel(self, start_date, end_date, group_id=None):
        """
        Экспорт расписания в Excel с красивым форматированием
        """
        try:
            logger.info(f"Starting Excel export for {start_date} - {end_date}")
            
            from .models import Group, Schedule
            
            # Получаем данные
            query = self.db.query(Schedule).filter(
                Schedule.date >= start_date,
                Schedule.date <= end_date
            ).order_by(Schedule.date, Schedule.start_time)
            
            if group_id:
                query = query.filter(Schedule.group_id == group_id)
            
            schedules = query.all()
            
            if not schedules:
                # Создаем пустой DataFrame с сообщением
                df = pd.DataFrame([{'Информация': 'Нет данных за выбранный период'}])
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Расписание', index=False)
                output.seek(0)
                return output
            
            # Группируем по группам
            groups_dict = {}
            for s in schedules:
                if s.group.name not in groups_dict:
                    groups_dict[s.group.name] = []
                groups_dict[s.group.name].append(s)
            
            # Создаем Excel файл
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Для каждой группы создаем отдельный лист
                for group_name, group_schedules in groups_dict.items():
                    self._create_group_excel_sheet(writer, group_name, group_schedules, start_date, end_date)
            
            output.seek(0)
            logger.info(f"Excel file created, size: {len(output.getvalue())} bytes")
            return output
            
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e
    
    def _create_group_excel_sheet(self, writer, group_name, schedules, start_date, end_date):
        """Создание листа Excel для одной группы"""
        
        # Группируем по датам
        days_data = {}
        for s in schedules:
            date_str = s.date.strftime('%d.%m.%Y')
            if date_str not in days_data:
                day_name = self.days_ru[s.date.weekday()] if s.date.weekday() < len(self.days_ru) else ''
                days_data[date_str] = {
                    'date': date_str,
                    'day': day_name,
                    'pairs': {i: [] for i in range(1, 7)}
                }
            
            pair_num = 1
            for i, slot in enumerate(self.time_slots, 1):
                if s.start_time in slot["time"]:
                    pair_num = i
                    break
            
            days_data[date_str]['pairs'][pair_num].append({
                'subject': s.subject.name if s.subject else '',
                'teacher': s.teacher.name if s.teacher else '',
                'room': s.room
            })
        
        sorted_dates = sorted(days_data.keys(), key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
        
        # Создаем данные для таблицы
        data = []
        header = ['№', 'Время']
        for date_str in sorted_dates:
            day_info = days_data[date_str]
            header.append(f"{day_info['day']}\n{date_str}")
        data.append(header)
        
        for pair_num in range(1, 7):
            row = [pair_num, self.time_slots[pair_num-1]["time"]]
            for date_str in sorted_dates:
                items = days_data[date_str]['pairs'][pair_num]
                if items:
                    cell_text = []
                    for item in items:
                        cell_text.append(f"{item['subject']}\n{item['teacher']}\n{item['room']}")
                    row.append('\n'.join(cell_text))
                else:
                    row.append('—')
            data.append(row)
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        sheet_name = group_name[:31]
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Заголовок группы
        worksheet.merge_cells(f'A1:{get_column_letter(len(header))}1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f"Расписание группы {group_name}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 30
        
        # Информация о периоде
        worksheet.merge_cells(f'A2:{get_column_letter(len(header))}2')
        period_cell = worksheet.cell(row=2, column=1)
        period_cell.value = f"Период: с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
        period_cell.font = Font(size=12, italic=True, color="4B5563")
        period_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Форматирование заголовков таблицы
        for col_num, value in enumerate(header, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Форматирование данных
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num in range(4, len(data) + 3):
            for col_num in range(1, len(header) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if (row_num - 3) % 2 == 0:
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        # Настройка ширины колонок
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 15
        
        for col_num in range(3, len(header) + 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    def export_to_pdf(self, start_date, end_date, group_id=None):
        """
        Экспорт расписания в PDF с красивым форматированием
        """
        try:
            logger.info(f"Starting PDF export for {start_date} - {end_date}")
            
            from .models import Group, Schedule
            
            schedules = self.db.query(Schedule).filter(
                Schedule.date >= start_date,
                Schedule.date <= end_date
            ).order_by(Schedule.date, Schedule.start_time).all()
            
            if group_id:
                schedules = [s for s in schedules if s.group_id == group_id]
            
            if not schedules:
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=A4,
                    rightMargin=2*cm,
                    leftMargin=2*cm,
                    topMargin=2*cm,
                    bottomMargin=2*cm
                )
                
                story = []
                styles = getSampleStyleSheet()
                
                story.append(Paragraph("Расписание занятий", styles['Title']))
                story.append(Spacer(1, 1*cm))
                story.append(Paragraph("Нет данных за выбранный период", styles['Normal']))
                
                doc.build(story)
                buffer.seek(0)
                return buffer
            
            # Группируем по группам
            groups_dict = {}
            for s in schedules:
                if s.group.name not in groups_dict:
                    groups_dict[s.group.name] = []
                groups_dict[s.group.name].append(s)
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                rightMargin=1*cm,
                leftMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1*cm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=1,
                spaceAfter=20,
                textColor=colors.HexColor('#4F46E5')
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,
                spaceAfter=30,
                textColor=colors.HexColor('#6B7280')
            )
            
            story.append(Paragraph("Расписание занятий", title_style))
            period_str = f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
            story.append(Paragraph(period_str, subtitle_style))
            
            for group_name, group_schedules in groups_dict.items():
                story.append(Paragraph(f"Группа: {group_name}", styles['Heading2']))
                
                days_data = {}
                for s in group_schedules:
                    date_str = s.date.strftime('%d.%m.%Y')
                    if date_str not in days_data:
                        day_name = self.days_ru[s.date.weekday()] if s.date.weekday() < len(self.days_ru) else ''
                        days_data[date_str] = {
                            'date': date_str,
                            'day': day_name,
                            'pairs': {i: [] for i in range(1, 7)}
                        }
                    
                    pair_num = 1
                    for i, slot in enumerate(self.time_slots, 1):
                        if s.start_time in slot["time"]:
                            pair_num = i
                            break
                    
                    days_data[date_str]['pairs'][pair_num].append({
                        'subject': s.subject.name if s.subject else '',
                        'teacher': s.teacher.name if s.teacher else '',
                        'room': s.room
                    })
                
                sorted_dates = sorted(days_data.keys(), key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
                
                table_data = [['№', 'Время']]
                for date_str in sorted_dates:
                    day_info = days_data[date_str]
                    table_data[0].append(f"{day_info['day']}\n{date_str}")
                
                for pair_num in range(1, 7):
                    row = [str(pair_num), self.time_slots[pair_num-1]["time"]]
                    for date_str in sorted_dates:
                        items = days_data[date_str]['pairs'][pair_num]
                        if items:
                            cell_text = []
                            for item in items:
                                cell_text.append(f"{item['subject']}\n{item['teacher']}\n{item['room']}")
                            row.append('\n'.join(cell_text))
                        else:
                            row.append('—')
                    table_data.append(row)
                
                col_widths = [1.5*cm, 3*cm]
                for _ in range(len(sorted_dates)):
                    col_widths.append(3.5*cm)
                
                table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ])
                
                for i in range(1, len(table_data)):
                    if i % 2 == 0:
                        style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F9FAFB'))
                
                table.setStyle(style)
                story.append(table)
                story.append(Spacer(1, 0.5*cm))
            
            doc.build(story)
            buffer.seek(0)
            
            logger.info(f"PDF file created, size: {len(buffer.getvalue())} bytes")
            return buffer
            
        except Exception as e:
            logger.error(f"PDF export error: {str(e)}")
            import traceback
            traceback.print_exc()
            raise e